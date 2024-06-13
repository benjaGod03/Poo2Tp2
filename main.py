import customtkinter
import networkx as nx
from customtkinter import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import openpyxl as op
from PIL import Image


class Grafo:
    def __init__(self, nombrePag):
        ws = op.load_workbook("Archivo.xlsx")
        wb = ws[nombrePag]
        self.g = nx.Graph()
        agregar = [wb.cell(row=i, column=1).value for i in range(1, wb.max_row +1)]
        if agregar[wb.max_row-1] is None:
            agregar.remove(agregar[wb.max_row-1])
        self.g.add_nodes_from(agregar)
        j = 1
        while wb['A{}'.format(j)].value is not None:
            values = [wb.cell(row=j, column=c).value for c in range(1, wb.max_column + 1)]
            nodo = values[0]
            p = 1
            while values[p] != None:
                if self.g.has_edge(nodo, values[p]) == False :
                    if self.g.has_node(values[p]) == True :
                        self.g.add_edge(nodo, values[p], amistad=values[p+1])
                if (p+2 >= values.__len__()):
                    break
                p += 2
            j += 1
    def getGraph(self):
        return self.g

a1 = Grafo('Grafo 1')
a2 = Grafo('Grafo 2')
g1 = a1.getGraph()
g2 = a2.getGraph()


grado_de_amistad = {"amigo personal" : 3, "conocido" : 2, "compañero" : 1}

def asignar_valor_amistad(G):
    for u, v, data in G.edges(data=True):
        data["valor"] = grado_de_amistad[data["amistad"]]

asignar_valor_amistad(g1)
asignar_valor_amistad(g2)

def distancia_de_amistad (G, source, target):
    try:
        distancia = nx.shortest_path_length(G, source=source, target=target, weight="valor")
        return distancia
    except nx.NetworkXNoPath:
        return float("inf")


def mostrar_grafo(frame_grafo, G):
    fig, ax = plt.subplots(figsize=(8, 6))

    pos = nx.spring_layout(G, seed=42)


    node_colors = [G.degree(n) for n in G.nodes()]
    node_sizes = [700 + 100 * G.degree(n) for n in G.nodes()]

    nx.draw(G, pos, with_labels=True, node_color=node_colors, node_size=node_sizes,
            cmap=plt.cm.viridis, font_weight='bold', edge_color='gray', ax=ax)


    labels = nx.get_edge_attributes(G, 'valor')
    nx.draw_networkx_edge_labels(G, pos, edge_labels=labels, font_color='red', ax=ax)

    canvas = FigureCanvasTkAgg(fig, master=frame_grafo)
    canvas.draw()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)



def calcular_distancia(G,n1,n2):
    nombre1 = n1.get().capitalize()
    nombre2 = n2.get().capitalize()
    distancia = distancia_de_amistad(G, nombre1, nombre2)
    resultado=f"La distancia de amistad entre {nombre1} y {nombre2} es: {distancia}"
    return resultado



class Window:
    def __init__(self):
        my_image = customtkinter.CTkImage(light_image=Image.open('friends.png'),
                                          size=(200, 200))
        self.app = CTk(fg_color="#174873")
        self.app.geometry("550x375")
        frame_prin = CTkFrame(self.app, fg_color="#F2F2F2", border_color="#1E0342")
        frame_prin.pack(pady=20, padx=20, fill="both", expand=True)

        label_titl = CTkLabel(frame_prin, text="Friends app", font=("Georgia", 21), text_color="#F2F2F2", fg_color="#BF2424", anchor='center',corner_radius=8)
        label_titl.pack(pady=5)
        label_sub = CTkLabel(frame_prin, text="Seleccione el Grafo de amistad:",font=("Georgia", 18), text_color="#F2F2F2", fg_color="#BF2424", corner_radius=8)
        label_sub.pack(pady=5)

        foto_label = customtkinter.CTkLabel(frame_prin, text="", image=my_image)
        foto_label.pack(pady=10)

        self.combobox = CTkComboBox(master=frame_prin, values=["Grafo 1", "Grafo 2"], corner_radius=2, fg_color="#591212",
                                     border_color="#BF2424", dropdown_hover_color="#BF2424", dropdown_fg_color="#295FA6", button_color="#BF2424", dropdown_text_color="#000000", command=self.opciones)
        self.combobox.place(relx=0.5, rely=0.5, anchor="center")

        self.app.mainloop()

    def opciones(self, value):
        if value == "Grafo 1":
            self.mostrar2(g1, "Grafo 1")
        else:
            self.mostrar2(g2, "Grafo 2")

    def mostrar2(self,g, nombre):
        ventana = CTkToplevel(fg_color="#174873")
        ventana.geometry("800x700")
        ventana.title(nombre)
        my_image = customtkinter.CTkImage(light_image=Image.open('friends.png'),
                                          size=(50, 50))
        frame_entrada = CTkFrame(ventana, fg_color="#F2F2F2", border_color="#1E0342")
        frame_entrada.pack(pady=20, padx=20, fill="both", expand=True)
        label_nombre1 = CTkLabel(frame_entrada, text="Introduce un nombre:", font=("Georgia", 15), text_color="#BF2424")
        label_nombre1.pack(pady=5)
        entry_nombre1 = CTkEntry(frame_entrada, width=200, fg_color="#174873")
        entry_nombre1.pack(pady=5)
        label_nombre2 = CTkLabel(frame_entrada, text="Introduce otro nombre:", font=("Georgia", 15), text_color="#BF2424")
        label_nombre2.pack(pady=5)
        entry_nombre2 = CTkEntry(frame_entrada, width=200,fg_color="#174873")
        entry_nombre2.pack(pady=5)
        button_calcular = CTkButton(frame_entrada, text="Calcular distancia",fg_color="#BF2424", command=lambda: label_resultado.configure(text=calcular_distancia(g,entry_nombre1,entry_nombre2)))
        button_calcular.pack(pady=10)
        foto_label = customtkinter.CTkLabel(frame_entrada, text="", image=my_image)
        foto_label.pack(pady=10)
        label_resultado = CTkLabel(frame_entrada, text="",text_color="#BF2424")
        label_resultado.pack(pady=10)
        frame_grafo = CTkFrame(ventana)
        frame_grafo.pack(pady=20, padx=20, fill="both", expand=True)
        mostrar_grafo(frame_grafo, g)


window = Window()
